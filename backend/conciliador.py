"""
Lógica principal de conciliación contable entre dos archivos de Excel.

Lee Excel con openpyxl y compara por fecha y monto (sin pandas).
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .bancos_extractos import BANCOS_EXTRACTOS
from .normalizador import normalizar_concepto, normalizar_fecha, normalizar_monto


@dataclass
class ColumnConfig:
    fecha: str
    concepto: str
    monto: str
    monto_creditos: Optional[str] = None
    monto_debitos: Optional[str] = None


def _inferir_columnas(columnas: List[str]) -> ColumnConfig:
    cols_lower = {c.strip().lower() if c else "": c for c in columnas if c is not None}
    cols_lower = {k: v for k, v in cols_lower.items() if k}

    def buscar(posibles: List[str]) -> str:
        for candidato in posibles:
            if candidato in cols_lower:
                return cols_lower[candidato]
        raise ValueError(
            f"No se encontró ninguna de las columnas requeridas: {', '.join(posibles)}"
        )

    def buscar_opt(posibles: List[str]) -> Optional[str]:
        for candidato in posibles:
            if candidato in cols_lower:
                return cols_lower[candidato]
        return None

    fecha = buscar([
        "fecha", "fecha extracto", "fecha gasto", "fecha_contable", "fecha_acreditacion",
        "fecha_emision", "f_extracto", "f_gasto", "f_contable", "fecha ato", "fecha comp", "fecha valor",
    ])
    concepto = buscar([
        "concepto", "descripcion", "detalle", "movimiento",
        "concepto extracto", "concepto gasto", "concepto contable", "nombre_cuenta",
        "orden_pago", "nro_movimiento", "nro_deposito",
    ])
    creditos = buscar_opt(["creditos", "crédito", "credito"])
    debitos = buscar_opt(["debitos", "débito", "debito"])
    if creditos is not None and debitos is not None:
        return ColumnConfig(
            fecha=fecha, concepto=concepto, monto=creditos,
            monto_creditos=creditos, monto_debitos=debitos,
        )
    monto = buscar([
        "monto", "importe", "valor", "monto extracto", "monto gasto", "monto contable",
        "neto", "saldo", "importe_debe", "importe_haber",
    ])
    return ColumnConfig(fecha=fecha, concepto=concepto, monto=monto)


def _column_config_para_banco(banco_id: str, columnas: List[str]) -> ColumnConfig:
    """
    Obtiene la ColumnConfig para un banco de extractos usando los nombres
    de columnas definidos para ese banco. Los headers del Excel deben coincidir.
    """
    if banco_id not in BANCOS_EXTRACTOS:
        raise ValueError(f"Banco desconocido: '{banco_id}'.")
    conf_banco = BANCOS_EXTRACTOS[banco_id]
    cols_lower = {c.strip().lower() if c else "": c for c in columnas if c is not None}
    cols_lower = {k: v for k, v in cols_lower.items() if k}

    def buscar(posibles: List[str]) -> str:
        for candidato in posibles:
            if candidato in cols_lower:
                return cols_lower[candidato]
        raise ValueError(
            f"No se encontró ninguna de las columnas requeridas para este banco: {', '.join(posibles)}"
        )

    def buscar_opt(posibles: List[str]) -> Optional[str]:
        for candidato in posibles:
            if candidato in cols_lower:
                return cols_lower[candidato]
        return None

    fecha = buscar(conf_banco["fecha"])
    concepto = buscar(conf_banco["concepto"])
    creditos = buscar_opt(conf_banco.get("creditos", []))
    debitos = buscar_opt(conf_banco.get("debitos", []))
    if creditos is not None and debitos is not None:
        return ColumnConfig(
            fecha=fecha, concepto=concepto, monto=creditos,
            monto_creditos=creditos, monto_debitos=debitos,
        )
    monto = buscar(conf_banco["monto"])
    return ColumnConfig(fecha=fecha, concepto=concepto, monto=monto)


def leer_excel_en_memoria(
    file_bytes: bytes,
    banco_extractos_id: Optional[str] = None,
    sheet_index: Optional[int] = None,
    max_rows: int = 200_000,
) -> List[Dict[str, Any]]:
    """
    Lee un archivo Excel desde bytes y devuelve una lista de diccionarios (una fila por dict).
    Usa por defecto la primera hoja. Si se indica sheet_index (1-based), usa esa hoja.
    Prueba varias filas como encabezado (0..11) para soportar informes con títulos.
    Si banco_extractos_id está definido, usa la configuración de columnas de ese banco (solo para extractos).
    Limita la cantidad de filas procesadas con max_rows para evitar consumir demasiada memoria.
    """
    buffer = BytesIO(file_bytes)
    ultimo_error: Exception | None = None

    try:
        wb = load_workbook(buffer, read_only=True, data_only=True)
    except Exception as exc:  # errores de formato/corrupción se manejan más arriba
        raise exc

    try:
        if sheet_index is not None:
            if sheet_index < 1 or sheet_index > len(wb.sheetnames):
                raise ValueError(
                    "El número de hoja seleccionado no es válido para este archivo."
                )
            ws = wb[wb.sheetnames[sheet_index - 1]]
        else:
            ws = wb.active

        # Construimos una lista de filas visibles (no ocultas por filtros)
        # y no completamente vacías. Esto respeta lo que ve el usuario en Excel.
        rows: List[List[Any]] = []
        visible_count = 0
        # Usamos enumerate para obtener el índice de fila, evitando depender de cell.row,
        # que no siempre existe en EmptyCell cuando read_only=True.
        for row_index, row in enumerate(ws.iter_rows(), start=1):
            if not row:
                continue
            # En modo read_only (ReadOnlyWorksheet) no siempre existe row_dimensions.
            if hasattr(ws, "row_dimensions"):
                row_dim = ws.row_dimensions.get(row_index)
                if row_dim is not None and getattr(row_dim, "hidden", False):
                    # Fila oculta por filtro u otra razón: se ignora
                    continue

            values = [cell.value for cell in row]
            if not any(v is not None and str(v).strip() for v in values):
                # Fila completamente vacía o con solo blancos: ignorar
                continue

            rows.append(values)
            visible_count += 1
            if visible_count > max_rows:
                raise ValueError(
                    f"El archivo Excel tiene demasiadas filas visibles ({visible_count}). "
                    f"Reducí el tamaño (máximo permitido: {max_rows})."
                )

        if not rows:
            raise ValueError("El archivo Excel no contiene filas de datos visibles.")

        max_header_row = min(12, len(rows))
        for header_row in range(max_header_row):
            try:
                if len(rows) <= header_row:
                    continue

                headers = [str(c).strip() if c is not None else "" for c in rows[header_row]]
                config = (
                    _column_config_para_banco(banco_extractos_id, headers)
                    if banco_extractos_id
                    else _inferir_columnas(headers)
                )

                data_rows = rows[header_row + 1 :]
                out: List[Dict[str, Any]] = []
                for row_values in data_rows:
                    if not any(v is not None and str(v).strip() for v in row_values):
                        continue
                    fila = dict(zip(headers, row_values)) if row_values else {}
                    out.append(fila)
                if out:
                    return out
            except ValueError as e:
                ultimo_error = e
                continue
    finally:
        wb.close()

    raise ValueError(
        ultimo_error.args[0] if ultimo_error else "El archivo Excel no contiene filas de datos válidos."
    )


def _preparar_filas(
    filas: List[Dict[str, Any]],
    column_config: Optional[ColumnConfig] = None,
    banco_extractos_id: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Aplica normalizaciones y filtra filas inválidas. Si se pasa column_config o banco_extractos_id, se usa en lugar de inferir."""
    if not filas:
        return []
    headers = list(filas[0].keys())
    if column_config is not None:
        config = column_config
    elif banco_extractos_id is not None:
        config = _column_config_para_banco(banco_extractos_id, headers)
    else:
        config = _inferir_columnas(headers)

    resultado = []
    for idx, fila in enumerate(filas):
        fecha_val = fila.get(config.fecha)
        concepto_val = fila.get(config.concepto)
        fecha_norm = normalizar_fecha(fecha_val)
        concepto_norm = normalizar_concepto(concepto_val)
        if config.monto_creditos is not None and config.monto_debitos is not None:
            cr = normalizar_monto(fila.get(config.monto_creditos)) or 0.0
            db = normalizar_monto(fila.get(config.monto_debitos)) or 0.0
            monto_val = cr - db
            monto_norm = round(monto_val, 2)
        else:
            monto_val = fila.get(config.monto)
            monto_norm = normalizar_monto(monto_val)

        if fecha_norm is None or monto_norm is None:
            continue
        if not (concepto_norm or concepto_norm.strip()):
            continue

        resultado.append({
            "id": idx,
            "fecha": fecha_val,
            "concepto": concepto_val,
            "monto": monto_val,
            "fecha_norm": fecha_norm,
            "concepto_norm": concepto_norm,
            "monto_norm": monto_norm,
        })
    return resultado


def preparar_filas_extractos(
    filas: List[Dict[str, Any]],
    banco_extractos_id: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Prepara filas de extractos (normaliza fecha, concepto, monto) para conciliación.
    Si banco_extractos_id es None, se infieren las columnas.
    """
    return _preparar_filas(filas, banco_extractos_id=banco_extractos_id)


def comparar_movimientos(
    extractos_filas: List[Dict[str, Any]],
    contable_filas: List[Dict[str, Any]],
    extractos_banco_id: Optional[str] = None,
    extractos_preparados: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """
    Compara movimientos por fecha y monto. Devuelve los que no tienen contraparte.
    Si extractos_banco_id está definido, se usa la config de columnas de ese banco para los extractos.
    Si extractos_preparados está definido, se usa esa lista en lugar de preparar extractos_filas
    (útil cuando se combinan hoja de extracto + hoja de cheques diferidos).
    """
    if extractos_preparados is not None:
        extractos = extractos_preparados
    else:
        extractos = _preparar_filas(extractos_filas, banco_extractos_id=extractos_banco_id)
    contable = _preparar_filas(contable_filas)

    # Índice (fecha_norm, monto_norm) -> lista de ids contable para emparejar 1 a 1.
    # Así la búsqueda es O(1) por extracto en lugar de O(m), evitando timeout en Render.
    contable_por_clave: Dict[Tuple[Any, Any], List[int]] = defaultdict(list)
    for row in contable:
        key = (row["fecha_norm"], row["monto_norm"])
        contable_por_clave[key].append(row["id"])

    usados_extracto: set[int] = set()
    usados_contable: set[int] = set()

    for ext in extractos:
        id_e = ext["id"]
        if id_e in usados_extracto:
            continue
        key = (ext["fecha_norm"], ext["monto_norm"])
        candidatos = contable_por_clave.get(key)
        if candidatos:
            cont_id = candidatos.pop()
            usados_extracto.add(id_e)
            usados_contable.add(cont_id)

    solo_en_extractos: List[Dict[str, Any]] = []
    solo_en_contable: List[Dict[str, Any]] = []

    for row in extractos:
        if row["id"] in usados_extracto:
            continue
        fn = row["fecha_norm"]
        fecha_str = fn.strftime("%d/%m/%Y") if fn else ""
        solo_en_extractos.append({
            "fecha": str(fecha_str),
            "monto": float(row["monto_norm"]),
            "descripcion": str(row["concepto"]) if row["concepto"] is not None else "",
        })

    for row in contable:
        if row["id"] in usados_contable:
            continue
        fn = row["fecha_norm"]
        fecha_str = fn.strftime("%d/%m/%Y") if fn else ""
        solo_en_contable.append({
            "fecha": str(fecha_str),
            "monto": float(row["monto_norm"]),
            "descripcion": str(row["concepto"]) if row["concepto"] is not None else "",
        })

    return {
        "solo_en_extractos": solo_en_extractos,
        "solo_en_contable": solo_en_contable,
        "resumen": {
            "total_extractos": len(extractos),
            "total_contable": len(contable),
            "coincidencias": len(usados_extracto),
            "diferentes_extractos": len(solo_en_extractos),
            "diferentes_contable": len(solo_en_contable),
        },
    }


def _obtener_valor_por_indice(fila: Dict[str, Any], indice: int) -> Any:
    """
    Devuelve el valor de una fila (dict) por índice de columna, respetando
    el orden original de las columnas del Excel (orden de creación del dict).
    Si el índice es inválido, devuelve None.
    """
    if indice < 0:
        return None
    try:
        return list(fila.values())[indice]
    except IndexError:
        return None


def comparar_por_columnas(
    extractos_filas: List[Dict[str, Any]],
    contable_filas: List[Dict[str, Any]],
    modo_comparacion: str,
) -> Dict[str, Any]:
    """
    Compara los archivos usando columnas específicas de fecha, concepto y monto
    (sin inferencia automática), según el modo seleccionado.

    Reglas pedidas:
    - Fecha:
        * Extracto: columna "Fecha"
        * Contable: columna "FECHA ATO"
      Solo se consideran filas que tengan fecha válida.
    - Concepto:
        * Extracto: columna "Movimiento"
        * Contable: columna "DETALLE"
      Se usa solo para mostrar la descripción en la salida.
    - Monto: depende del modo de comparación (índices de columna, 0-based)
        * "extracto_D_vs_contable_I":
            - Extracto: columna D (índice 3)
            - Contable: columna I (índice 8)
        * "contable_H_vs_extracto_E":
            - Extracto: columna E (índice 4)
            - Contable: columna H (índice 7)
      El emparejamiento se hace por (fecha_norm, monto_norm).
    """
    if modo_comparacion == "extracto_D_vs_contable_I":
        idx_monto_ext = 3  # Columna D
        idx_monto_cont = 8  # Columna I
    elif modo_comparacion == "contable_H_vs_extracto_E":
        idx_monto_ext = 4  # Columna E
        idx_monto_cont = 7  # Columna H
    else:
        raise ValueError(f"Modo de comparación no soportado: {modo_comparacion}")

    # Preparar filas de extracto y contable con las columnas pedidas.
    def preparar_filas_especial(
        filas: List[Dict[str, Any]],
        fecha_col: str,
        concepto_col: str,
        idx_monto: int,
    ) -> List[Dict[str, Any]]:
        if not filas:
            return []
        resultado: List[Dict[str, Any]] = []
        for idx, fila in enumerate(filas):
            fecha_val = fila.get(fecha_col)
            fecha_norm = normalizar_fecha(fecha_val)
            if fecha_norm is None:
                # Solo mostrar lo que tenga fecha
                continue
            concepto_val = fila.get(concepto_col)
            concepto_norm = normalizar_concepto(concepto_val)
            monto_val = _obtener_valor_por_indice(fila, idx_monto)
            monto_norm = normalizar_monto(monto_val)
            if monto_norm is None:
                continue
            resultado.append(
                {
                    "id": idx,
                    "fecha": fecha_val,
                    "concepto": concepto_val,
                    "monto": monto_val,
                    "fecha_norm": fecha_norm,
                    "concepto_norm": concepto_norm,
                    "monto_norm": monto_norm,
                }
            )
        return resultado

    extractos = preparar_filas_especial(
        extractos_filas,
        fecha_col="Fecha",
        concepto_col="Movimiento",
        idx_monto=idx_monto_ext,
    )
    contable = preparar_filas_especial(
        contable_filas,
        fecha_col="FECHA ATO",
        concepto_col="DETALLE",
        idx_monto=idx_monto_cont,
    )

    if not extractos or not contable:
        return {
            "solo_en_extractos": [],
            "solo_en_contable": [],
            "resumen": {
                "total_extractos": len(extractos),
                "total_contable": len(contable),
                "coincidencias": 0,
                "diferentes_extractos": 0,
                "diferentes_contable": 0,
            },
        }

    # Índice (fecha_norm, monto_norm) -> lista de ids contable para emparejar 1 a 1.
    contable_por_clave: Dict[Tuple[Any, Any], List[int]] = defaultdict(list)
    for row in contable:
        key = (row["fecha_norm"], row["monto_norm"])
        contable_por_clave[key].append(row["id"])

    usados_extracto: set[int] = set()
    usados_contable: set[int] = set()

    for ext in extractos:
        id_e = ext["id"]
        if id_e in usados_extracto:
            continue
        key = (ext["fecha_norm"], ext["monto_norm"])
        candidatos = contable_por_clave.get(key)
        if candidatos:
            cont_id = candidatos.pop()
            usados_extracto.add(id_e)
            usados_contable.add(cont_id)

    solo_en_extractos: List[Dict[str, Any]] = []
    solo_en_contable: List[Dict[str, Any]] = []

    for row in extractos:
        if row["id"] in usados_extracto:
            continue
        fn = row["fecha_norm"]
        fecha_str = fn.strftime("%d/%m/%Y") if fn else ""
        solo_en_extractos.append(
            {
                "fecha": str(fecha_str),
                "monto": float(row["monto_norm"]),
                "descripcion": str(row["concepto"]) if row["concepto"] is not None else "",
            }
        )

    for row in contable:
        if row["id"] in usados_contable:
            continue
        fn = row["fecha_norm"]
        fecha_str = fn.strftime("%d/%m/%Y") if fn else ""
        solo_en_contable.append(
            {
                "fecha": str(fecha_str),
                "monto": float(row["monto_norm"]),
                "descripcion": str(row["concepto"]) if row["concepto"] is not None else "",
            }
        )

    return {
        "solo_en_extractos": solo_en_extractos,
        "solo_en_contable": solo_en_contable,
        "resumen": {
            "total_extractos": len(extractos),
            "total_contable": len(contable),
            "coincidencias": len(usados_extracto),
            "diferentes_extractos": len(solo_en_extractos),
            "diferentes_contable": len(solo_en_contable),
        },
    }
