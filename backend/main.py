"""
Punto de entrada del backend FastAPI para el conciliador contable.

Expone un endpoint POST /conciliar que recibe dos archivos Excel
en formato multipart/form-data. Compara por fecha y monto y genera
un Excel con las diferencias en outputs/.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
import logging
import time

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from . import config
from .bancos_extractos import BANCOS_EXTRACTOS, get_bancos_lista
from .conciliador import (
    comparar_movimientos,
    comparar_por_columnas,
    leer_excel_en_memoria,
    preparar_filas_extractos,
)
from .schemas import ErrorResponse


logger = logging.getLogger("conciliador")

MAX_UPLOAD_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB por archivo


app = FastAPI(
    title="Conciliador Contable",
    description="API para conciliación contable basada en archivos Excel.",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(HTTPException)
async def http_exception_handler(_, exc: HTTPException):
    return JSONResponse(
        status_code=exc.status_code,
        content=ErrorResponse(detail=str(exc.detail)).dict(),
    )


@app.post(
    "/conciliar",
    responses={
        200: {"description": "JSON con los movimientos que difieren entre ambos Excels."},
        400: {"model": ErrorResponse, "description": "Error de validación o entrada."},
        500: {"model": ErrorResponse, "description": "Error interno del servidor."},
    },
)
def conciliar_endpoint(
    extractos_file: UploadFile = File(..., description="Archivo Excel con los extractos."),
    contable_file: UploadFile = File(..., description="Archivo Excel con el registro contable."),
    banco_extractos: str = Form(..., description="Banco del extracto (santander o provincia)."),
    extractos_hoja_index: int = Form(
        ...,
        description="Número de hoja (1-based) del archivo de extractos a usar como extracto principal.",
    ),
    contable_hoja_index: int = Form(
        ...,
        description="Número de hoja (1-based) del archivo contable a analizar.",
    ),
    cheques_diferidos_hoja_index: int | None = Form(
        None,
        description="Número de hoja (1-based) del archivo de extractos donde están los cheques diferidos (opcional).",
    ),
    modo_comparacion: str = Form(
        "fecha_monto",
        description=(
            "Modo de comparación: 'fecha_monto' (por defecto), "
            "'extracto_D_vs_contable_I' o 'contable_H_vs_extracto_E'."
        ),
    ),
):
    inicio_total = time.perf_counter()
    logger.info(
        "Iniciando conciliación: banco=%s, extractos_hoja=%s, contable_hoja=%s, cheques_hoja=%s",
        banco_extractos,
        extractos_hoja_index,
        contable_hoja_index,
        cheques_diferidos_hoja_index,
    )

    if not extractos_file or not contable_file:
        raise HTTPException(status_code=400, detail="Debe enviar ambos archivos: extractos y contable.")
    # Solo se valida que esté seleccionado un tipo de extracto en el combo; no se inspecciona el archivo.
    if not banco_extractos or banco_extractos not in BANCOS_EXTRACTOS:
        raise HTTPException(
            status_code=400,
            detail="Seleccioná el tipo de extracto (Santander, Banco Provincia o Banco Nación) en el combo.",
        )

    if extractos_hoja_index < 1 or contable_hoja_index < 1:
        raise HTTPException(
            status_code=400,
            detail="Los números de hoja deben ser mayores o iguales a 1.",
        )

    try:
        # Lectura de archivos en memoria (lado síncrono: este endpoint corre en threadpool)
        extractos_bytes = extractos_file.file.read()
        contable_bytes = contable_file.file.read()

        if not extractos_bytes:
            raise HTTPException(status_code=400, detail="El archivo de extractos está vacío.")
        if not contable_bytes:
            raise HTTPException(status_code=400, detail="El archivo contable está vacío.")

        logger.info(
            "Archivos recibidos: extractos=%d bytes, contable=%d bytes",
            len(extractos_bytes),
            len(contable_bytes),
        )

        try:
            t0 = time.perf_counter()
            # Hoja principal de extractos (siempre obligatoria)
            extractos_principal = leer_excel_en_memoria(
                extractos_bytes,
                banco_extractos_id=banco_extractos,
                sheet_index=extractos_hoja_index,
            )
            logger.info(
                "Lectura de extractos completada: %d filas en %.3f s",
                len(extractos_principal),
                time.perf_counter() - t0,
            )

            # Hoja de cheques diferidos (opcional). Se lee sin configuración de banco
            # porque suele tener otra estructura de columnas.
            extractos_combinados = None
            if cheques_diferidos_hoja_index is not None:
                t1 = time.perf_counter()
                extractos_cheques = leer_excel_en_memoria(
                    extractos_bytes,
                    banco_extractos_id=None,
                    sheet_index=cheques_diferidos_hoja_index,
                )
                logger.info(
                    "Lectura de cheques diferidos completada: %d filas en %.3f s",
                    len(extractos_cheques),
                    time.perf_counter() - t1,
                )
                prep_principal = preparar_filas_extractos(
                    extractos_principal, banco_extractos_id=banco_extractos
                )
                prep_cheques = preparar_filas_extractos(
                    extractos_cheques, banco_extractos_id=None
                )
                extractos_combinados = prep_principal + prep_cheques
                for i, row in enumerate(extractos_combinados):
                    row["id"] = i
                extractos_filas = None
            else:
                extractos_filas = extractos_principal

            t2 = time.perf_counter()
            contable_filas = leer_excel_en_memoria(
                contable_bytes,
                banco_extractos_id=None,
                sheet_index=contable_hoja_index,
            )
            logger.info(
                "Lectura de contable completada: %d filas en %.3f s",
                len(contable_filas),
                time.perf_counter() - t2,
            )
        except ValueError as e:
            logger.warning("Error de validación al leer Excel: %s", e)
            raise HTTPException(status_code=400, detail=str(e))
        except HTTPException:
            raise
        except Exception:
            logger.exception("Error inesperado leyendo los archivos Excel")
            raise HTTPException(
                status_code=400,
                detail="No fue posible leer uno de los archivos Excel. Verificá que el formato sea válido (.xlsx) "
                "y que las hojas seleccionadas correspondan al tipo de archivo esperado.",
            )

        t3 = time.perf_counter()
        if modo_comparacion == "fecha_monto":
            if extractos_combinados is not None:
                resultado = comparar_movimientos(
                    [], contable_filas, extractos_preparados=extractos_combinados
                )
            else:
                resultado = comparar_movimientos(
                    extractos_filas, contable_filas, extractos_banco_id=banco_extractos
                )
        else:
            # Modos especiales por columnas específicas; se trabaja solo con la hoja principal.
            try:
                resultado = comparar_por_columnas(
                    extractos_principal,
                    contable_filas,
                    modo_comparacion=modo_comparacion,
                )
            except ValueError as e:
                logger.warning("Error en modo de comparación por columnas: %s", e)
                raise HTTPException(status_code=400, detail=str(e))
        logger.info(
            "Comparación de movimientos completada en %.3f s", time.perf_counter() - t3
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"comparacion_{timestamp}.xlsx"
        output_path = config.OUTPUTS_DIR / filename

        font_cuerpo = Font(name="Calibri", size=14)
        font_titulo = Font(name="Calibri", size=14, bold=True)

        def formatear_hoja(ws):
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font_cuerpo
            for cell in ws[1]:
                cell.font = font_titulo
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                if col:
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 1, 80)

        t4 = time.perf_counter()
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Solo en extractos"
        ws1.append(["Fecha", "Monto", "Descripción"])
        for r in resultado["solo_en_extractos"]:
            ws1.append([r["fecha"], r["monto"], r["descripcion"]])
        formatear_hoja(ws1)

        ws2 = wb.create_sheet("Solo en contable")
        ws2.append(["Fecha", "Monto", "Descripción"])
        for r in resultado["solo_en_contable"]:
            ws2.append([r["fecha"], r["monto"], r["descripcion"]])
        formatear_hoja(ws2)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        output_path.write_bytes(buffer.read())
        logger.info(
            "Generación de Excel de salida completada en %.3f s",
            time.perf_counter() - t4,
        )

        resultado["excel_filename"] = filename
        resultado["modo_comparacion"] = modo_comparacion
        logger.info("Conciliación finalizada en %.3f s", time.perf_counter() - inicio_total)
        return resultado

    except HTTPException:
        # Ya encapsula un mensaje claro para el frontend; solo lo registramos.
        logger.warning("Conciliación abortada por HTTPException", exc_info=True)
        raise
    except Exception as e:
        logger.exception("Error interno no controlado durante la conciliación")
        raise HTTPException(
            status_code=500,
            detail="Ocurrió un error interno al comparar los archivos. Intentá nuevamente más tarde. "
            f"Detalle técnico: {e}",
        )


@app.get("/descargar/{filename}")
async def descargar_excel(filename: str):
    if not filename.endswith(".xlsx") or ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Nombre de archivo no válido.")
    path = config.OUTPUTS_DIR / filename
    if not path.exists():
        raise HTTPException(status_code=404, detail="El archivo no existe o ya fue eliminado.")
    return FileResponse(
        path=str(path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@app.get("/bancos")
async def listar_bancos():
    """Devuelve la lista de bancos disponibles para el archivo de extractos."""
    return {"bancos": get_bancos_lista()}


@app.post(
    "/info_extracto",
    responses={
        200: {"description": "Información básica del archivo de extractos (cantidad de hojas)."},
        400: {"model": ErrorResponse, "description": "Error de validación o entrada."},
    },
)
async def info_extracto(
    extractos_file: UploadFile = File(..., description="Archivo Excel con los extractos."),
):
    """Devuelve la cantidad de hojas del Excel de extractos y sus nombres."""
    if not extractos_file:
        raise HTTPException(status_code=400, detail="Debe enviar un archivo de extractos.")

    contenido = await extractos_file.read()
    if not contenido:
        raise HTTPException(status_code=400, detail="El archivo de extractos está vacío.")

    if len(contenido) > MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(
            status_code=400,
            detail="El archivo de extractos es demasiado grande para inspeccionarlo desde la web. "
            "Reducí el tamaño del Excel e intentá de nuevo.",
        )

    try:
        buffer = BytesIO(contenido)
        wb = load_workbook(buffer, read_only=True, data_only=True)
        nombres = list(wb.sheetnames)
        wb.close()
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="No fue posible leer el archivo de extractos. Verifique que el formato sea válido (.xlsx).",
        )

    return {
        "total_hojas": len(nombres),
        "hojas": [
            {"indice": idx + 1, "nombre": nombre}
            for idx, nombre in enumerate(nombres)
        ],
    }


@app.post(
    "/info_contable",
    responses={
        200: {"description": "Información básica del archivo contable (cantidad de hojas)."},
        400: {"model": ErrorResponse, "description": "Error de validación o entrada."},
    },
)
async def info_contable(
    contable_file: UploadFile = File(..., description="Archivo Excel contable."),
):
    """Devuelve la cantidad de hojas del Excel contable y sus nombres."""
    if not contable_file:
        raise HTTPException(status_code=400, detail="Debe enviar un archivo contable.")

    contenido = await contable_file.read()
    if not contenido:
        raise HTTPException(status_code=400, detail="El archivo contable está vacío.")

    try:
        buffer = BytesIO(contenido)
        wb = load_workbook(buffer, read_only=True, data_only=True)
        nombres = list(wb.sheetnames)
        wb.close()
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="No fue posible leer el archivo contable. Verifique que el formato sea válido (.xlsx).",
        )

    return {
        "total_hojas": len(nombres),
        "hojas": [
            {"indice": idx + 1, "nombre": nombre}
            for idx, nombre in enumerate(nombres)
        ],
    }


@app.get("/health")
async def health_check():
    return {"status": "ok"}


# Sirve el frontend estático (para despliegue en Render; en local opcional).
if config.FRONTEND_DIR.exists():
    app.mount("/", StaticFiles(directory=str(config.FRONTEND_DIR), html=True), name="frontend")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("backend.main:app", host="0.0.0.0", port=config.PORT, reload=True)
